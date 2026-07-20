import json
from pathlib import Path
from openpyxl import load_workbook

INPUT_FILES = [Path('data.xlsx'), Path('data/data.xlsx')]
OUTPUT_DIR = Path('data')
CHUNK_COUNT = 4

input_path = next((p for p in INPUT_FILES if p.exists()), None)
if input_path is None:
    raise FileNotFoundError('data.xlsx not found in project root or data/ directory')

wb = load_workbook(input_path, data_only=True, read_only=True)
ws = wb['النظام الجديد'] if 'النظام الجديد' in wb.sheetnames else wb.worksheets[0]
rows = list(ws.iter_rows(values_only=True))
if not rows:
    raise ValueError('Excel sheet is empty')

header = [str(cell) if cell is not None else '' for cell in rows[0]]
data_rows = []
for r in rows[1:]:
    if not r or r[0] is None:
        continue
    row = []
    for cell in r[:5]:
        if isinstance(cell, str):
            row.append(cell)
        elif isinstance(cell, (int, float)):
            row.append(cell)
        elif cell is None:
            row.append(None)
        else:
            row.append(str(cell))
    if len(row) < 5:
        row += [None] * (5 - len(row))
    data_rows.append(row)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
chunk_size = max(1, (len(data_rows) + CHUNK_COUNT - 1) // CHUNK_COUNT)
for i in range(CHUNK_COUNT):
    start = i * chunk_size
    end = start + chunk_size
    part_rows = data_rows[start:end]
    if not part_rows:
        continue
    part_file = OUTPUT_DIR / f'results_part{i+1}.js'
    output_rows = [header] + part_rows
    part_file.write_text(
        'window.RESULTS_DATA_ROWS = ' + json.dumps(output_rows, ensure_ascii=False, separators=(',', ':')) + ';\n',
        encoding='utf-8'
    )
    print(f'Wrote {part_file} rows={len(part_rows)}')

print('Done. Generated', len(list(OUTPUT_DIR.glob('results_part*.js'))), 'part files.')
