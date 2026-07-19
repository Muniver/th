import json
from pathlib import Path
from openpyxl import load_workbook

p = Path('data.xlsx')
print('exists', p.exists(), 'size', p.stat().st_size if p.exists() else None)
wb = load_workbook(p, data_only=True)

result = {}
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    result[ws.title] = rows

out_path = Path('excel_data.json')
with out_path.open('w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print('wrote', out_path)
print('sheets', wb.sheetnames)
for name, rows in result.items():
    print('sheet', name, 'rows', len(rows))
    for row in rows[:8]:
        print(row)
    print('---')
